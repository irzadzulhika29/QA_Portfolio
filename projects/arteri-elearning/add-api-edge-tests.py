"""Tambah 10 negative + 5 edge test case ke Arteri API sheet."""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = "/home/ubuntu/testing-hub/projects/arteri-elearning/test-cases.xlsx"

# Styling
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
PRIORITY_HIGH = PatternFill("solid", fgColor="FECACA")
PRIORITY_MED = PatternFill("solid", fgColor="FED7AA")
PRIORITY_LOW = PatternFill("solid", fgColor="DCFCE7")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================
# NEW API TEST CASES: 10 NEGATIVE (security-focused) + 5 EDGE
# ============================================================
NEW_API = [
    # ===== NEGATIVE (Security) =====
    ("API-AUTH-014", "/auth/me", "GET", "Negative",
     "JWT replay attack (reuse token setelah logout)",
     "User logout, simpan token lama",
     "(with logged-out token)",
     '401 Unauthorized, message: "Token revoked" (server harus track revoked tokens)',
     "High"),
    ("API-AUTH-015", "/auth/register", "POST", "Negative",
     "Mass assignment attack (kirim field role=admin)",
     "—",
     '{"name": "x", "email": "x@test.com", "password": "Valid123!", "role": "admin", "is_admin": true}',
     '201 Created, TAPI role yang tersimpan = "student" (default), field tambahan diabaikan/di-strip',
     "High"),
    ("API-STU-014", "/courses/browse", "GET", "Negative",
     "SQL injection di search query",
     "—",
     "?search=test%27%20OR%20%271%27%3D%271",
     '200 OK dengan list kosong atau error, TIDAK return semua data (sanitized)',
     "High"),
    ("API-STU-015", "/courses/browse", "GET", "Negative",
     "SQL injection dengan UNION SELECT",
     "—",
     "?search=test%27%20UNION%20SELECT%20*%20FROM%20users--",
     '400 Bad Request atau 200 dengan list kosong, BUKAN data user terexpose',
     "High"),
    ("API-CBT-016", "/upcoming-exam", "GET", "Negative",
     "CORS bypass attempt (Origin: attacker.com)",
     "—",
     "Header Origin: https://attacker.com",
     "Response TIDAK punya Access-Control-Allow-Origin: https://attacker.com (CORS strict)",
     "High"),
    ("API-CBT-017", "/student/exams-attempt/:id/answers", "POST", "Negative",
     "Negative index di question (manipulate question_id)",
     "—",
     '{"question_id": -1, "selected_option_id": ["a"]}',
     '400 Bad Request, "Invalid question_id"',
     "Medium"),
    ("API-CBT-018", "/student/exams-attempt/:id/answers", "POST", "Negative",
     "Mass assignment (kirim is_correct, score field)",
     "—",
     '{"question_id": "q1", "selected_option_id": ["a"], "is_correct": true, "score": 100}',
     '201 Created (atau 200), TAPI score dan is_correct di-strip (server-side compute)',
     "High"),
    ("API-TCH-013", "/admin/users", "GET", "Negative",
     "Privilege escalation (teacher non-admin coba akses admin endpoint)",
     "Login as teacher biasa (non-admin)",
     "(with teacher token)",
     '403 Forbidden, "Admin access required"',
     "High"),
    ("API-TCH-014", "/auth/login-elearning", "POST", "Negative",
     "Timing attack (response time beda untuk valid vs invalid user)",
     "—",
     "Compare timing: email valid + random password vs random email",
     "Response time konsisten (< 100ms difference) untuk valid dan invalid user (timing-safe comparison)",
     "Medium"),
    ("API-EXM-014", "/teacher/courses/:id/exam", "POST", "Negative",
     "Mass assignment (kirim is_published=true di body create)",
     "—",
     '{"title": "x", "duration": 60, "due_date": "2026-12-31", "is_published": true, "created_by": "fake_id"}',
     '201 Created, TAPI is_published default false (draft), created_by di-strip (server ambil dari token)',
     "High"),

    # ===== EDGE (Boundary) =====
    ("API-AUTH-016", "/auth/login-elearning", "POST", "Edge",
     "Empty body (no Content-Length, no body)",
     "—",
     "(empty body, no Content-Type)",
     '400 Bad Request atau 411 Length Required, TIDAK crash server',
     "Medium"),
    ("API-STU-016", "/courses/browse", "GET", "Edge",
     "Max size query string (> 8KB)",
     "—",
     "?search=<8000 character random string>",
     '200 OK dengan list kosong (atau 414 URI Too Long), server tidak crash',
     "Medium"),
    ("API-STU-017", "/student/notes", "POST", "Edge",
     "Unicode/emoji di input",
     "—",
     '{"topic_id": 1, "content": "测试 🎓 العربية 한국어 🦄"}',
     '201 Created, content tersimpan dengan benar (UTF-8 support verified)',
     "Medium"),
    ("API-CBT-019", "/teacher/exams/:id/questions", "POST", "Edge",
     "Very long question text (10,000 karakter)",
     "—",
     '{"question_text": "<10K random chars>", "question_type": "single", "options": [...], "correct": "a"}',
     '201 Created atau 413 Payload Too Large (kalau ada max length, return 400 dengan pesan jelas)',
     "Medium"),
    ("API-TCH-015", "/student/courses/99999999999999999999", "GET", "Edge",
     "Integer overflow pada path parameter",
     "—",
     "/student/courses/99999999999999999999",
     '400 Bad Request (invalid ID) atau 404 Not Found, TIDAK crash dengan NumberFormatError',
     "Medium"),
]

# ============================================================
# ADD TO XLSX
# ============================================================
wb = load_workbook(XLSX)
ws_api = wb["API Test Cases"]

next_row = ws_api.max_row + 1

for tc in NEW_API:
    row_data = list(tc) + ["Belum Dieksekusi", "", ""]
    for col_idx, val in enumerate(row_data, start=1):
        ws_api.cell(row=next_row, column=col_idx, value=val)

    for col_idx in range(1, len(row_data) + 1):
        cell = ws_api.cell(row=next_row, column=col_idx)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER

    priority = tc[8]
    if priority == "High":
        ws_api.cell(row=next_row, column=9).fill = PRIORITY_HIGH
    elif priority == "Medium":
        ws_api.cell(row=next_row, column=9).fill = PRIORITY_MED
    elif priority == "Low":
        ws_api.cell(row=next_row, column=9).fill = PRIORITY_LOW

    next_row += 1

# Update Summary
ws_summary = wb["Summary"]
api_total = ws_api.max_row - 1

# Find rows
for r in range(10, 16):
    cell_a = ws_summary.cell(row=r, column=1).value
    if cell_a == "API Test Cases":
        ws_summary.cell(row=r, column=2, value=api_total)
    elif cell_a == "Total":
        # Hitung ulang total = UI + API
        ui_total = None
        for r2 in range(10, 16):
            if ws_summary.cell(row=r2, column=1).value == "UI Test Cases":
                ui_total = ws_summary.cell(row=r2, column=2).value
                break
        if ui_total:
            ws_summary.cell(row=r, column=2, value=ui_total + api_total)

# Update by-endpoint count
ep_added = {}
for tc in NEW_API:
    endpoint = tc[1].split("/")[1] if "/" in tc[1] else tc[1]
    ep_added[endpoint] = ep_added.get(endpoint, 0) + 1

for r in range(20, 40):
    cell_a = ws_summary.cell(row=r, column=1).value
    if cell_a is None:
        continue
    cell_a_stripped = cell_a.strip() if isinstance(cell_a, str) else ""
    for ep in ep_added:
        if cell_a_stripped == f"/{ep}":
            old = ws_summary.cell(row=r, column=2).value
            try:
                old_int = int(old) if old is not None else 0
            except (TypeError, ValueError):
                old_int = 0
            ws_summary.cell(row=r, column=2, value=old_int + ep_added[ep])
            break

# Save
wb.save(XLSX)
size = os.path.getsize(XLSX)
print(f"OK Arteri XLSX updated")
print(f"  File: {XLSX}")
print(f"  Size: {size:,} bytes")
print(f"  API Test Cases total: {api_total} (added {len(NEW_API)})")
print(f"\nAdded breakdown:")
print(f"  Negative (security): 10")
print(f"  Edge (boundary):     5")
print(f"  Total added:         15")
print(f"\nBy endpoint:")
for ep, c in sorted(ep_added.items()):
    print(f"  /{ep}: +{c}")
