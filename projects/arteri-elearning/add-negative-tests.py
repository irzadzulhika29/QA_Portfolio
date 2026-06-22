"""Tambah negative test case ke Arteri XLSX existing."""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/home/ubuntu/testing-hub/projects/arteri-elearning/test-cases.xlsx"

# Styling (sama dengan existing)
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
PRIORITY_HIGH = PatternFill("solid", fgColor="FECACA")
PRIORITY_MED = PatternFill("solid", fgColor="FED7AA")
PRIORITY_LOW = PatternFill("solid", fgColor="DCFCE7")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================
# NEW NEGATIVE TEST CASES (UI)
# Format: (ID, Modul, Tipe, Item, Test Case, Pre-condition, Steps, Expected, Priority)
# IDs lanjut dari existing: AUTH-011+, CBT-028+, STU-015+, ADM-017+
# ============================================================
NEGATIVE_UI = [
    # ===== AUTH (7 baru) =====
    ("AUTH-011", "Auth", "Negative", "Login",
     "Login dengan email mengandung SQL injection",
     "—",
     "1. Buka /login\n2. Isi email: ' OR 1=1--\n3. Isi password apa saja\n4. Klik Masuk",
     "Login ditolak, error 'Format email tidak valid' (input disanitasi)", "High"),
    ("AUTH-012", "Auth", "Negative", "Login",
     "Login dengan password mengandung XSS script",
     "—",
     "1. Buka /login\n2. Isi email valid\n3. Isi password: <script>alert(1)</script>\n4. Klik Masuk",
     "Login gagal (kredensial salah), input di-escape, no script execution", "High"),
    ("AUTH-013", "Auth", "Negative", "Login",
     "Login dengan email lebih dari 255 karakter",
     "—",
     "1. Isi email 300 karakter (long string)\n2. Submit",
     "Validasi max length, error 'Email maksimal 255 karakter'", "Medium"),
    ("AUTH-014", "Auth", "Negative", "JWT",
     "JWT dengan signature dimodifikasi (tampered)",
     "User login",
     "1. Login\n2. Modify signature di localStorage JWT\n3. Refresh page",
     "401 Unauthorized, redirect ke /login (server reject signature)", "High"),
    ("AUTH-015", "Auth", "Negative", "JWT",
     "JWT dengan alg 'none' (algorithm confusion attack)",
     "—",
     "1. Set localStorage JWT dengan header alg=none\n2. Buka protected route",
     "401 Unauthorized, server reject unsigned token", "High"),
    ("AUTH-016", "Auth", "Negative", "Brute Force",
     "Login gagal 10x berturut-turut (brute force)",
     "—",
     "1. Buka /login\n2. Submit password salah 10x dalam 5 menit",
     "Setelah 5x: rate limit aktif, response 429 Too Many Requests", "High"),
    ("AUTH-017", "Auth", "Negative", "Cookie",
     "Cookie access_token dimanipulasi / dihapus",
     "User login",
     "1. Login\n2. Hapus cookie access_token via DevTools\n3. Refresh",
     "Middleware tidak bisa decode, redirect ke /login", "Medium"),

    # ===== CBT (10 baru - CRITICAL) =====
    ("CBT-028", "CBT", "Negative", "Timer",
     "Timer mundur ke nilai negatif",
     "—",
     "1. Manipulasi timer state ke -1 (via DevTools)",
     "Timer auto-submit, atau UI block invalid state", "High"),
    ("CBT-029", "CBT", "Negative", "Timer",
     "Timer window blur (pindah window)，会不会 pause timer",
     "Exam in progress",
     "1. Pindah window ke aplikasi lain (Alt+Tab)\n2. Tunggu 5 menit\n3. Kembali",
     "Timer TETAP jalan (server-side validation), tidak di-pause oleh blur", "High"),
    ("CBT-030", "CBT", "Negative", "Tab Switch",
     "Tab switch > 3x (multiple violations)",
     "Sudah 2 lives hilang",
     "1. Switch tab 5x berturut-turut cepat",
     "Semua switch ke-record, lives jadi 0, auto-disqualify", "High"),
    ("CBT-031", "CBT", "Negative", "Tab Switch",
     "Tab switch dengan DevTools disable event listener",
     "—",
     "1. Buka DevTools, remove window.blur listener\n2. Switch tab",
     "Backend tetap catat tab switch (server-side validation), lives decrement", "High"),
    ("CBT-032", "CBT", "Negative", "Auto Save",
     "Auto-save dengan payload super besar (DoS attempt)",
     "—",
     "1. Inject localStorage dengan 100MB data\n2. Trigger auto-save",
     "Auto-save gagal, tidak crash app, error logged, retry dengan snapshot kecil", "Medium"),
    ("CBT-033", "CBT", "Negative", "Answer State",
     "Manipulasi Redux state via DevTools (cheating)",
     "Exam in progress",
     "1. Buka DevTools\n2. Inject jawaban langsung ke Redux state\n3. Submit",
     "Backend verifikasi jawaban dengan attempt_id, hanya jawaban tersimpan yang dihitung", "High"),
    ("CBT-034", "CBT", "Negative", "Resume",
     "Resume attempt yang sudah disubmit (race condition)",
     "—",
     "1. Submit exam di tab A\n2. Buka tab B, refresh (attempt_id sama)",
     "Tab B: exam sudah disubmit, redirect ke result page", "Medium"),
    ("CBT-035", "CBT", "Negative", "Question",
     "Soal dengan question_type tidak dikenal dari backend",
     "—",
     "1. Backend return soal dengan question_type='unknown_type'",
     "Component fallback ke error state, exam tidak crash, soal diskip atau di-flag", "Medium"),
    ("CBT-036", "CBT", "Negative", "Question",
     "Soal dengan options kosong (corrupt data)",
     "—",
     "1. Backend return soal single_choice tanpa options",
     "Soal tampil dengan pesan 'Soal tidak valid', tidak bisa dijawab, flag otomatis", "Medium"),
    ("CBT-037", "CBT", "Negative", "Matching",
     "Soal matching dengan pair tidak lengkap (hanya left atau right)",
     "—",
     "1. Backend return soal matching dengan options side=left saja",
     "Service normalize & fallback, soal tetap renderable", "Medium"),

    # ===== STUDENT (5 baru) =====
    ("STU-015", "Student", "Negative", "Enroll",
     "Enroll ke course yang tidak ada",
     "—",
     "1. POST /student/enroll dengan course_id=99999 (invalid)",
     "404 Not Found, error message 'Course tidak ditemukan'", "Medium"),
    ("STU-016", "Student", "Negative", "Enroll",
     "Double enroll (submit form 2x cepat)",
     "—",
     "1. Klik Enroll 2x cepat (double-click)",
     "Hanya 1 enroll yang terbuat, idempotent", "Medium"),
    ("STU-017", "Student", "Negative", "Quiz",
     "Submit quiz tanpa jawab semua soal",
     "—",
     "1. Mulai quiz\n2. Submit tanpa jawab semua",
     "Submit tetap sukses, soal unanswered = 0 (atau warning)", "Medium"),
    ("STU-018", "Student", "Negative", "Quiz",
     "Start quiz yang sudah pernah di-attempt (resume vs new)",
     "—",
     "1. Quiz pernah di-start, attempt_id tersimpan\n2. Klik Mulai Quiz lagi",
     "Backend return attempt yang sama (resume), tidak duplicate", "Medium"),
    ("STU-019", "Student", "Negative", "Notifications",
     "Mark as read notifikasi yang sudah dihapus (race)",
     "—",
     "1. Hapus notifikasi X\n2. Klik notifikasi X dari cache (sebelum refresh)",
     "404 atau silent ignore, no error popup", "Low"),

    # ===== ADMIN (5 baru) =====
    ("ADM-017", "Admin", "Negative", "CRUD",
     "Create course dengan field kosong (title, description)",
     "Teacher login",
     "1. Buka form Tambah Course\n2. Submit tanpa isi title",
     "Validasi form gagal, error 'Title wajib diisi'", "Medium"),
    ("ADM-018", "Admin", "Negative", "Permission",
     "Teacher non-admin akses endpoint admin-only",
     "Login as teacher (non-admin)",
     "1. Kirim POST ke /admin/users (admin only)",
     "403 Forbidden", "High"),
    ("ADM-019", "Admin", "Negative", "Exam",
     "Create exam dengan due_date di masa lalu",
     "—",
     "1. Form Tambah Exam, isi due_date kemarin",
     "Validasi gagal, error 'Due date harus di masa depan'", "Medium"),
    ("ADM-020", "Admin", "Negative", "Bulk Upload",
     "Bulk upload dengan CSV format salah",
     "—",
     "1. Upload CSV dengan kolom missing / typo",
     "Error report: list baris yang gagal + alasan, partial success atau full fail", "Medium"),
    ("ADM-021", "Admin", "Negative", "Cheating",
     "Cheating report untuk exam yang tidak ada",
     "—",
     "1. Buka /dashboard-admin/cheating-report?exam_id=99999",
     "Empty state, pesan 'Belum ada data cheating'", "Low"),
]

# ============================================================
# NEW NEGATIVE API TEST CASES
# Format: (ID, Endpoint, Method, Tipe, Test Case, Pre-condition, Request, Expected, Priority)
# IDs lanjut: API-AUTH-009+, API-STU-010+, API-CBT-010+, API-TCH-009+, API-EXM-011+
# ============================================================
NEGATIVE_API = [
    # ===== AUTH API (5 baru) =====
    ("API-AUTH-009", "/auth/login-elearning", "POST", "Negative",
     "Login dengan email mengandung SQL injection",
     "—",
     '{"email": "admin@test.com\\\" OR 1=1--", "password": "x"}',
     '400 Bad Request (input disanitasi) atau 401 Unauthorized, BUKAN 200 dengan data leak',
     "High"),
    ("API-AUTH-010", "/auth/login-elearning", "POST", "Negative",
     "Login tanpa body",
     "—",
     "{}",
     '400 Bad Request, validation error',
     "Medium"),
    ("API-AUTH-011", "/auth/login-elearning", "POST", "Negative",
     "Login dengan Content-Type bukan JSON",
     "—",
     "Content-Type: text/plain, body=raw",
     '415 Unsupported Media Type atau 400 Bad Request',
     "Low"),
    ("API-AUTH-012", "/auth/refresh", "POST", "Negative",
     "Refresh dengan token invalid",
     "—",
     '{"refresh_token": "invalid"}',
     '401 Unauthorized',
     "Medium"),
    ("API-AUTH-013", "/auth/refresh", "POST", "Negative",
     "Refresh dengan refresh_token expired",
     "—",
     '{"refresh_token": "<expired>"}',
     '401 Unauthorized, message: "Refresh token expired"',
     "Medium"),

    # ===== STUDENT API (4 baru) =====
    ("API-STU-010", "/student/enroll", "POST", "Negative",
     "Enroll tanpa auth (token)",
     "—",
     '{"course_id": 1}',
     '401 Unauthorized',
     "High"),
    ("API-STU-011", "/student/enroll", "POST", "Negative",
     "Enroll dengan course_id bukan number",
     "—",
     '{"course_id": "abc"}',
     '400 Bad Request, validation error',
     "Medium"),
    ("API-STU-012", "/student/courses/abc", "GET", "Negative",
     "Get course dengan ID invalid (non-numeric)",
     "—",
     "/student/courses/abc",
     '400 Bad Request atau 404 Not Found',
     "Medium"),
    ("API-STU-013", "/student/content/123", "GET", "Negative",
     "Get content yang tidak ada",
     "—",
     "/student/content/99999",
     '404 Not Found',
     "Medium"),

    # ===== CBT API (6 baru - CRITICAL) =====
    ("API-CBT-010", "/student/exams-attempt/:id/answers", "POST", "Negative",
     "Save answer setelah exam disubmit (race condition)",
     "Exam sudah disubmit",
     '{"question_id": "q1", "selected_option_id": ["a"]}',
     '409 Conflict, "Cannot modify submitted exam"',
     "High"),
    ("API-CBT-011", "/student/exams-attempt/:id/tab-switch", "POST", "Negative",
     "Tab switch spam (50x request dalam 1 detik)",
     "—",
     "50x POST dalam 1 detik",
     'Rate limit 429 setelah threshold, atau lives langsung 0',
     "High"),
    ("API-CBT-012", "/student/exams-attempt/:id/submit", "POST", "Negative",
     "Submit attempt orang lain (IDOR attempt)",
     "Login as student A",
     'Submit attempt_id milik student B',
     '403 Forbidden, "Cannot access other user attempt"',
     "High"),
    ("API-CBT-013", "/student/exams-attempt/:id/answers", "POST", "Negative",
     "Save answer untuk question yang tidak ada di exam",
     "—",
     '{"question_id": "fake_id_999", "selected_option_id": ["a"]}',
     '400 Bad Request, "Invalid question"',
     "Medium"),
    ("API-CBT-014", "/student/exams-attempt/:id/answers", "POST", "Negative",
     "Save answer dengan selected_option_id kosong",
     "—",
     '{"question_id": "q1", "selected_option_id": []}',
     '400 Bad Request, validation error',
     "Medium"),
    ("API-CBT-015", "/upcoming-exam", "GET", "Negative",
     "Get upcoming exam tanpa auth",
     "—",
     "(no token)",
     '401 Unauthorized',
     "High"),

    # ===== TEACHER API (4 baru) =====
    ("API-TCH-009", "/teacher/add-course", "POST", "Negative",
     "Create course dengan title kosong",
     "Teacher login",
     '{"title": "", "description": "...", "subject": "math"}',
     '400 Bad Request, "Title required"',
     "Medium"),
    ("API-TCH-010", "/teacher/courses/9999", "PUT", "Negative",
     "Update course yang tidak ada",
     "—",
     '{"title": "Updated"}',
     '404 Not Found',
     "Medium"),
    ("API-TCH-011", "/teacher/courses/:id", "DELETE", "Negative",
     "Delete course yang punya student enrolled",
     "Course ada student",
     "(no body)",
     '409 Conflict, "Cannot delete course with active students" (atau soft delete + warning)',
     "Medium"),
    ("API-TCH-012", "/teacher/add-course", "POST", "Negative",
     "Create course tanpa auth",
     "—",
     '{"title": "..."}',
     '401 Unauthorized',
     "High"),

    # ===== EXAM API (3 baru) =====
    ("API-EXM-011", "/teacher/courses/:id/exam", "POST", "Negative",
     "Create exam dengan duration <= 0",
     "—",
     '{"title": "UTS", "duration": 0, "due_date": "2026-12-31"}',
     '400 Bad Request, "Duration must be > 0"',
     "Medium"),
    ("API-EXM-012", "/teacher/courses/:id/exam", "POST", "Negative",
     "Create exam dengan due_date di masa lalu",
     "—",
     '{"title": "UTS", "duration": 60, "due_date": "2020-01-01"}',
     '400 Bad Request, "Due date must be in future"',
     "Medium"),
    ("API-EXM-013", "/teacher/exams/:id/questions", "POST", "Negative",
     "Add question tanpa options untuk single_choice",
     "—",
     '{"question_text": "Test", "question_type": "single", "options": []}',
     '400 Bad Request, "Options required for single choice"',
     "Medium"),
]

# ============================================================
# ADD TO XLSX
# ============================================================
wb = load_workbook(XLSX)

# Helper: add rows to sheet
def add_rows(ws, rows, header_styles=True):
    # Find next row
    next_row = ws.max_row + 1

    for tc in rows:
        # Append: list + ["Belum Dieksekusi", "", ""]
        row_data = list(tc) + ["Belum Dieksekusi", "", ""]
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx, value=val)

        # Style cells
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=next_row, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = BORDER

        # Priority color (col 9 for both sheets)
        priority = tc[8]
        if priority == "High":
            ws.cell(row=next_row, column=9).fill = PRIORITY_HIGH
        elif priority == "Medium":
            ws.cell(row=next_row, column=9).fill = PRIORITY_MED
        elif priority == "Low":
            ws.cell(row=next_row, column=9).fill = PRIORITY_LOW

        next_row += 1

    return next_row - 1  # last row added

# Add UI test cases
ws_ui = wb["Test Cases"]
last_ui = add_rows(ws_ui, NEGATIVE_UI)

# Add API test cases
ws_api = wb["API Test Cases"]
last_api = add_rows(ws_api, NEGATIVE_API)

# Update Summary
ws_summary = wb["Summary"]
# Find row "UI Test Cases" and "API Test Cases"
ui_count = len(NEGATIVE_UI) + (last_ui - len(NEGATIVE_UI) - 1)
api_count = len(NEGATIVE_API) + (last_api - len(NEGATIVE_API) - 1)

# Actually compute correctly
ui_total = ws_ui.max_row - 1  # minus header
api_total = ws_api.max_row - 1

# Update rows 11, 12, 13 (Test Inventory section)
for r in range(10, 14):
    cell_a = ws_summary.cell(row=r, column=1).value
    cell_b = ws_summary.cell(row=r, column=2).value
    if cell_a == "UI Test Cases":
        ws_summary.cell(row=r, column=2, value=ui_total)
    elif cell_a == "API Test Cases":
        ws_summary.cell(row=r, column=2, value=api_total)
    elif cell_a == "Total":
        ws_summary.cell(row=r, column=2, value=ui_total + api_total)

# Update per-modul (UI) - find section and add new entries
modul_added = {}
for tc in NEGATIVE_UI:
    modul = tc[1]
    modul_added.setdefault(modul, 0)
    modul_added[modul] += 1

# Find existing modul rows and update
for r in range(15, 35):
    cell_a = ws_summary.cell(row=r, column=1).value
    if cell_a is None:
        continue
    cell_a_stripped = cell_a.strip() if isinstance(cell_a, str) else ""
    for modul in modul_added:
        if cell_a_stripped == modul:
            old = ws_summary.cell(row=r, column=2).value or 0
            try:
                old_int = int(old)
            except (TypeError, ValueError):
                old_int = 0
            ws_summary.cell(row=r, column=2, value=old_int + modul_added[modul])
            break

# Update per-priority
prio_added = {"High": 0, "Medium": 0, "Low": 0}
for tc in NEGATIVE_UI:
    p = tc[8]
    if p in prio_added:
        prio_added[p] += 1

for r in range(20, 35):
    cell_a = ws_summary.cell(row=r, column=1).value
    if cell_a is None:
        continue
    cell_a_stripped = cell_a.strip() if isinstance(cell_a, str) else ""
    for p in prio_added:
        if cell_a_stripped == p:
            old = ws_summary.cell(row=r, column=2).value or 0
            ws_summary.cell(row=r, column=2, value=old + prio_added[p])
            break

# Save
wb.save(XLSX)
size = os.path.getsize(XLSX)
print(f"OK Arteri XLSX updated: {XLSX}")
print(f"  Size: {size:,} bytes")
print(f"  UI Test Cases: {ui_total} (added {len(NEGATIVE_UI)})")
print(f"  API Test Cases: {api_total} (added {len(NEGATIVE_API)})")
print(f"  Total added: {len(NEGATIVE_UI) + len(NEGATIVE_API)} negative test cases")
print()
print("Added by modul:")
for m, c in sorted(modul_added.items()):
    print(f"  {m}: +{c}")
print("Added by priority:")
for p, c in prio_added.items():
    print(f"  {p}: +{c}")
