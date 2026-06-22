"""
Generate XLSX Test Case untuk IFL Chapter Malang
Output: 2 sheets (Test Cases + API Test Cases) + Summary sheet
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = "/home/ubuntu/testing-hub/projects/ifl-chapter-malang/test-cases.xlsx"

# ============================================================
# STYLING
# ============================================================
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)

PRIORITY_HIGH = PatternFill("solid", fgColor="FECACA")    # red
PRIORITY_MED = PatternFill("solid", fgColor="FED7AA")     # orange
PRIORITY_LOW = PatternFill("solid", fgColor="DCFCE7")     # green

THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================
# TEST CASES — UI / Functional
# ============================================================
# Format: (ID, Modul, Tipe, Item, Test Case, Pre-condition, Steps, Expected, Priority)
TEST_CASES = [
    # ===== A. AUTHENTICATION =====
    ("AUTH-001", "Auth", "UI/Functional", "Login Form", "Login dengan email + password valid",
     "User sudah terdaftar di sistem",
     "1. Buka /masuk\n2. Isi email valid\n3. Isi password valid\n4. Klik tombol Masuk",
     "Login berhasil, redirect ke /admin/dashboards (admin) atau / (user)", "High"),
    ("AUTH-002", "Auth", "UI/Functional", "Login Form", "Login dengan email tidak valid (format salah)",
     "—",
     "1. Buka /masuk\n2. Isi email 'bukan-email'\n3. Isi password apa saja\n4. Klik Masuk",
     "Validasi form gagal, pesan error 'Format email tidak valid'", "Medium"),
    ("AUTH-003", "Auth", "UI/Functional", "Login Form", "Login dengan password kosong",
     "—",
     "1. Buka /masuk\n2. Isi email valid\n3. Kosongkan password\n4. Klik Masuk",
     "Validasi form gagal, pesan error 'Password wajib diisi'", "Medium"),
    ("AUTH-004", "Auth", "UI/Functional", "Login Form", "Login dengan kredensial salah",
     "User tidak ada / password salah",
     "1. Buka /masuk\n2. Isi email valid + password salah\n3. Klik Masuk",
     "Toast error 'Email atau password salah', tetap di halaman login", "High"),
    ("AUTH-005", "Auth", "UI/Functional", "Login Form", "Login dengan akun yang belum verified",
     "User terdaftar tapi belum klik link verifikasi",
     "1. Buka /masuk\n2. Login dengan akun unverified\n3. Klik Masuk",
     "Redirect ke /verifikasi, minta user verifikasi email dulu", "High"),
    ("AUTH-006", "Auth", "UI/Functional", "Register", "Register dengan email baru",
     "Email belum terdaftar",
     "1. Buka /daftar\n2. Isi form (nama, email, password, dll)\n3. Klik Daftar",
     "Akun dibuat, email verifikasi dikirim, redirect ke /verifikasi", "High"),
    ("AUTH-007", "Auth", "UI/Functional", "Register", "Register dengan email yang sudah ada",
     "Email sudah terdaftar",
     "1. Buka /daftar\n2. Isi email yang sudah ada\n3. Klik Daftar",
     "Toast error 'Email sudah terdaftar, silakan login'", "Medium"),
    ("AUTH-008", "Auth", "UI/Functional", "Register", "Password tidak memenuhi syarat (min 8 char, dll)",
     "—",
     "1. Buka /daftar\n2. Isi password 'abc' (terlalu pendek)\n3. Submit",
     "Validasi form gagal, pesan error 'Password minimal 8 karakter'", "Medium"),
    ("AUTH-009", "Auth", "UI/Functional", "Forgot Password", "Request reset password",
     "Email terdaftar",
     "1. Buka /lupa-kata-sandi\n2. Isi email\n3. Klik Kirim",
     "Email reset dikirim, pesan sukses 'Link reset telah dikirim ke email Anda'", "Medium"),
    ("AUTH-010", "Auth", "UI/Functional", "Verify", "Klik link verifikasi dari email",
     "User baru register, link ada di email",
     "1. Buka email verifikasi\n2. Klik link verifikasi",
     "Akun terverifikasi, redirect ke login dengan pesan 'Email berhasil diverifikasi'", "High"),
    ("AUTH-011", "Auth", "Integration", "Google OAuth", "Login dengan Google (sukses)",
     "User punya akun Google",
     "1. Buka /masuk\n2. Klik 'Login dengan Google'\n3. Pilih akun Google\n4. Authorize",
     "OAuth callback sukses, user ter-login, redirect ke home/dashboard sesuai role", "High"),
    ("AUTH-012", "Auth", "Integration", "Google OAuth", "Login dengan Google (user cancel)",
     "—",
     "1. Buka /masuk\n2. Klik 'Login dengan Google'\n3. Cancel di popup Google",
     "Kembali ke /masuk, tidak ada perubahan state", "Low"),
    ("AUTH-013", "Auth", "UI/Functional", "Logout", "Logout dari menu profile",
     "User sudah login",
     "1. Klik avatar/menu profile\n2. Klik Logout",
     "Token cleared, redirect ke /", "High"),
    ("AUTH-014", "Auth", "Security", "ProtectedToken", "Akses /chatbot/:slug tanpa login",
     "User belum login",
     "1. Tanpa login, langsung buka /chatbot/test",
     "Redirect ke /masuk, query param ?redirect=/chatbot/test", "High"),
    ("AUTH-015", "Auth", "Security", "ProtectedRoles", "User 'copywriter' akses /admin/dashboard/user (admin only)",
     "Login sebagai copywriter",
     "1. Login sebagai copywriter\n2. Buka /admin/dashboard/user",
     "Redirect ke /admin/dashboards atau 403 Forbidden", "High"),
    ("AUTH-016", "Auth", "Security", "ProtectedRoles", "User 'bismar' akses /admin/dashboard/blog (admin+copywriter only)",
     "Login sebagai bismar",
     "1. Login sebagai bismar\n2. Buka /admin/dashboard/blog",
     "Redirect ke /admin/dashboards atau 403 Forbidden", "High"),
    ("AUTH-017", "Auth", "Security", "JWT", "JWT expired — auto logout",
     "Login, tunggu token expired",
     "1. Login\n2. Modify token di localStorage jadi expired\n3. Refresh atau click link",
     "Token cleared, redirect ke /masuk", "High"),

    # ===== B. PUBLIC PAGES =====
    ("PUB-001", "Public", "UI/Visual", "Home", "Render halaman home dengan hero section",
     "—",
     "1. Buka /",
     "Hero tampil, CTA button visible, no console error", "Low"),
    ("PUB-002", "Public", "UI/Visual", "Home", "Home responsive di mobile (375px)",
     "—",
     "1. Buka / di viewport 375x667\n2. Inspect",
     "Layout responsive, hamburger menu visible, no horizontal scroll", "Medium"),
    ("PUB-003", "Public", "UI/Visual", "About", "Render halaman tentang-kami",
     "—",
     "1. Buka /tentang-kami",
     "Halaman tampil, sejarah & visi-misi visible", "Low"),
    ("PUB-004", "Public", "UI/Functional", "Event Listing", "Browse event listing publik",
     "Ada event aktif",
     "1. Buka /event",
     "List event tampil, card clickable ke detail", "Medium"),
    ("PUB-005", "Public", "UI/Functional", "Event Listing", "Filter event by category",
     "—",
     "1. Buka /event\n2. Pilih kategori",
     "List filter sesuai kategori", "Low"),
    ("PUB-006", "Public", "UI/Functional", "Project", "Browse project listing publik",
     "—",
     "1. Buka /project",
     "List project tampil, gambar & deskripsi visible", "Low"),
    ("PUB-007", "Public", "UI/Functional", "Blog Listing", "Browse artikel",
     "—",
     "1. Buka /artikel",
     "List artikel tampil dengan excerpt & cover image", "Medium"),
    ("PUB-008", "Public", "UI/Functional", "Blog Listing", "Pagination blog",
     "Ada > 10 artikel",
     "1. Buka /artikel\n2. Klik page 2",
     "Artikel page 2 tampil, URL update ?page=2", "Medium"),
    ("PUB-009", "Public", "UI/Functional", "Single Blog", "Baca artikel dengan rich text content",
     "—",
     "1. Buka /artikel/[slug]",
     "Konten CKEditor render dengan benar (heading, bold, list, image, code block)", "High"),
    ("PUB-010", "Public", "UI/Visual", "Department Pages", "Akses /nondept, /cdsi, /bismar, /dpp, /hrd, /comper",
     "—",
     "1. Buka masing-masing page department",
     "Semua page department tampil dengan benar", "Low"),
    ("PUB-011", "Public", "UI/Visual", "404", "Akses URL yang tidak ada",
     "—",
     "1. Buka /url-tidak-ada",
     "Halaman 404 Not Found tampil dengan link kembali ke home", "Low"),
    ("PUB-012", "Public", "UI/Visual", "Coming Soon", "Akses /coming-soon",
     "—",
     "1. Buka /coming-soon",
     "Halaman Coming Soon tampil", "Low"),

    # ===== C. DONATION + MIDTRANS (HIGHEST PRIORITY) =====
    ("DON-001", "Donasi", "UI/Functional", "Browse Donasi", "Lihat daftar campaign donasi",
     "—",
     "1. Buka /donasi",
     "List campaign tampil dengan image, target, progress bar", "High"),
    ("DON-002", "Donasi", "UI/Functional", "Single Campaign", "Lihat detail campaign donasi",
     "—",
     "1. Buka /donasi/[slug]",
     "Detail campaign tampil: deskripsi, target, deadline, tombol Donasi", "High"),
    ("DON-003", "Donasi", "Integration", "Chatbot Start", "Mulai chatbot donasi (perlu login)",
     "User sudah login",
     "1. Buka /donasi/[slug]\n2. Klik 'Donasi Sekarang'\n3. (jika belum login) redirect ke /masuk\n4. Setelah login, kembali ke chatbot",
     "Chatbot terbuka di /chatbot/[slug]", "High"),
    ("DON-004", "Donasi", "UI/Functional", "Chatbot Composer", "Input nominal donasi",
     "Chatbot terbuka",
     "1. Klik input composer\n2. Ketik '100000'\n3. Submit",
     "Nominal tercatat, chatbot validasi (min, max, format)", "High"),
    ("DON-005", "Donasi", "UI/Functional", "Chatbot Validation", "Input nominal di bawah minimum",
     "—",
     "1. Input '500' (< minimum)",
     "Validation feedback: 'Nominal minimum Rp 10.000'", "Medium"),
    ("DON-006", "Donasi", "UI/Functional", "Chatbot Quick Reply", "Pilih nominal cepat",
     "—",
     "1. Klik quick reply 'Rp 50.000'",
     "Nominal terisi otomatis", "Medium"),
    ("DON-007", "Donasi", "UI/Functional", "Chatbot Summary", "Lihat summary card sebelum bayar",
     "Sudah isi nominal + data diri",
     "1. Lanjut ke step summary",
     "Summary card tampil: nominal, data donor, metode pembayaran, tombol Bayar", "High"),
    ("DON-008", "Donasi", "Integration", "Midtrans Snap", "Buka Midtrans Snap pop-up",
     "Di halaman summary, klik Bayar",
     "1. Klik 'Bayar Sekarang'",
     "Snap Midtrans terbuka, daftar metode pembayaran (GoPay, OVO, DANA, QRIS, VA, CC)", "High"),
    ("DON-009", "Donasi", "Integration", "Midtrans QRIS", "Bayar dengan QRIS",
     "Snap terbuka",
     "1. Pilih QRIS\n2. Scan QR dengan e-wallet\n3. Konfirmasi di e-wallet",
     "Payment success, redirect ke /donasi/pembayaran/sukses", "High"),
    ("DON-010", "Donasi", "Integration", "Midtrans VA", "Bayar dengan Virtual Account",
     "Snap terbuka",
     "1. Pilih Bank Transfer (BCA/Mandiri/dll)\n2. Dapat nomor VA\n3. Bayar via mobile banking",
     "Payment success setelah VA dibayar, redirect ke success page", "High"),
    ("DON-011", "Donasi", "Integration", "Midtrans E-Wallet", "Bayar dengan GoPay/OVO/DANA",
     "Snap terbuka",
     "1. Pilih e-wallet\n2. Konfirmasi di app e-wallet",
     "Payment success, redirect ke success page", "High"),
    ("DON-012", "Donasi", "UI/Functional", "Payment Success", "Lihat halaman sukses",
     "Payment berhasil",
     "1. Setelah payment success, otomatis redirect ke /donasi/pembayaran/sukses",
     "Halaman success tampil: 'Terima kasih atas donasi Anda!', nominal, transaction ID", "High"),
    ("DON-013", "Donasi", "Integration", "Payment Timeout", "Tutup Snap tanpa bayar (timeout)",
     "Snap terbuka, user diam 15+ menit",
     "1. Buka Snap\n2. Jangan bayar / close popup",
     "Session expired, user kembali ke summary, bisa coba lagi", "Medium"),
    ("DON-014", "Donasi", "Integration", "Payment Failure", "Bayar gagal (kartu ditolak)",
     "Snap terbuka, pakai test card declined",
     "1. Pilih Credit Card\n2. Masukkan test card declined (Visa 4811 1111 1111 1114 di sandbox)\n3. Submit",
     "Error message dari Midtrans, kembali ke summary, bisa coba lagi", "High"),
    ("DON-015", "Donasi", "UI/Functional", "Donation History", "Admin lihat history donasi",
     "Login sebagai admin/bismar",
     "1. Buka /admin/dashboard/donation/[slug]",
     "List donasi tampil: nama donor, nominal, tanggal, status", "High"),
    ("DON-016", "Donasi", "UI/Functional", "Chatbot Payment Info", "Lihat payment info di chatbot",
     "Chatbot di step payment",
     "1. Lihat payment info component",
     "Detail pembayaran tampil: nominal, biaya admin (jika ada), total", "Medium"),

    # ===== D. BLOG + CKEDITOR =====
    ("BLG-001", "Blog", "UI/Functional", "Search Blog", "Search artikel by keyword",
     "—",
     "1. Buka /artikel\n2. Ketik keyword di search\n3. Submit",
     "List filter sesuai keyword", "Medium"),
    ("BLG-002", "Blog", "UI/Functional", "Blog Categories", "Filter blog by kategori",
     "—",
     "1. Buka /artikel\n2. Pilih kategori",
     "List filter sesuai kategori", "Medium"),
    ("BLG-003", "Blog", "UI/Functional", "Admin Add Blog", "Tambah blog baru (admin/copywriter)",
     "Login sebagai admin/copywriter",
     "1. Buka /admin/dashboard/blog/add\n2. Isi judul, kategori, konten (CKEditor)\n3. Upload cover (ImageKit)\n4. Klik Publish",
     "Blog tersimpan, toast sukses, redirect ke list", "High"),
    ("BLG-004", "Blog", "UI/Functional", "CKEditor", "Format rich text (bold, italic, heading, list)",
     "Di form Add Blog",
     "1. Klik toolbar Bold, Italic, Heading 1, Bullet List\n2. Ketik konten",
     "Format tampil sesuai", "High"),
    ("BLG-005", "Blog", "UI/Functional", "CKEditor", "Insert image di konten",
     "Di CKEditor",
     "1. Klik insert image\n2. Upload dari komputer / paste URL",
     "Image masuk konten, render dengan benar", "High"),
    ("BLG-006", "Blog", "UI/Functional", "CKEditor", "Insert link",
     "Di CKEditor",
     "1. Select text\n2. Klik insert link\n3. Isi URL",
     "Link ter-format", "Low"),
    ("BLG-007", "Blog", "UI/Functional", "ImageKit Upload", "Upload cover image",
     "Di form Add Blog",
     "1. Klik upload cover\n2. Pilih gambar (crop via react-cropper)\n3. Crop & save",
     "Image ter-upload ke ImageKit, URL ter-attach ke form", "High"),
    ("BLG-008", "Blog", "UI/Functional", "Admin Edit Blog", "Edit blog existing",
     "—",
     "1. Buka /admin/dashboard/blog\n2. Klik Edit di salah satu blog",
     "Form ter-isi dengan data existing, bisa diedit", "High"),
    ("BLG-009", "Blog", "Integration", "Admin Delete Blog", "Hapus blog",
     "—",
     "1. Buka /admin/dashboard/blog\n2. Klik Delete di salah satu blog\n3. Konfirmasi",
     "Blog terhapus, hilang dari list", "High"),
    ("BLG-010", "Blog", "UI/Functional", "Blog Categories Mgmt", "Admin manage kategori blog",
     "Login sebagai admin/copywriter",
     "1. Buka /admin/dashboard/category/blog\n2. Tambah/edit/hapus kategori",
     "Kategori ter-update", "Medium"),
    ("BLG-011", "Blog", "UI/Functional", "Blog Detail Admin", "Lihat detail blog di admin",
     "—",
     "1. Buka /admin/dashboard/blog/detail/[slug]",
     "Detail blog tampil lengkap", "Medium"),

    # ===== E. EVENT =====
    ("EVT-001", "Event", "UI/Functional", "Public Event", "Lihat detail event publik",
     "—",
     "1. Buka detail event dari /event",
     "Detail event tampil: tanggal, lokasi, deskripsi, tombol Daftar", "Medium"),
    ("EVT-002", "Event", "Integration", "Event Registration", "Daftar event",
     "User login",
     "1. Klik 'Daftar Event'\n2. Isi form pendaftaran\n3. Submit",
     "Pendaftaran sukses, dapat email konfirmasi", "High"),
    ("EVT-003", "Event", "UI/Functional", "Admin Event Management", "CRUD event",
     "Login sebagai admin",
     "1. Buka /admin/dashboard/event\n2. Create / Edit / Delete event",
     "Event ter-CRUD dengan benar", "High"),
    ("EVT-004", "Event", "UI/Functional", "Event Filter", "Filter event by date/category",
     "—",
     "1. Di /event, gunakan filter",
     "List filter sesuai filter", "Low"),
    ("EVT-005", "Event", "UI/Functional", "Event Chatbot", "Chatbot info event (jika ada)",
     "—",
     "1. Buka chatbot dari event",
     "Chatbot response dengan info event", "Low"),

    # ===== F. VOLUNTEER =====
    ("VOL-001", "Volunteer", "UI/Functional", "Volunteer Info", "Lihat info volunteer publik",
     "—",
     "1. Buka section volunteer di home atau /volunteer",
     "Info volunteer tampil", "Low"),
    ("VOL-002", "Volunteer", "UI/Functional", "Volunteer Registration", "Daftar jadi volunteer",
     "—",
     "1. Isi form volunteer\n2. Submit",
     "Pendaftaran volunteer terkirim", "High"),
    ("VOL-003", "Volunteer", "UI/Functional", "Admin Volunteer List", "Lihat list pendaftar volunteer",
     "Login sebagai admin",
     "1. Buka /admin/dashboard/volunteer",
     "List volunteer tampil", "Medium"),
    ("VOL-004", "Volunteer", "Integration", "Volunteer Approval", "Approve/reject volunteer",
     "—",
     "1. Di list volunteer, klik Approve / Reject",
     "Status volunteer ter-update, notifikasi email terkirim", "Medium"),
    ("VOL-005", "Volunteer", "Integration", "Volunteer Profile Match", "Match volunteer dengan kebutuhan",
     "—",
     "1. Sistem match profile volunteer dengan lowongan",
     "Rekomendasi volunteer sesuai skills", "Low"),

    # ===== G. PROFILE =====
    ("PROF-001", "Profile", "UI/Functional", "View Profile", "Lihat profil sendiri",
     "User login",
     "1. Klik avatar\n2. Pilih 'My Profile'",
     "Profile tampil: nama, foto, email, dll", "Medium"),
    ("PROF-002", "Profile", "UI/Functional", "Edit Profile", "Edit data profile",
     "—",
     "1. Di halaman profile, klik Edit\n2. Ubah data\n3. Save",
     "Data ter-update", "Medium"),
    ("PROF-003", "Profile", "UI/Functional", "Upload Profile Pic", "Upload & crop foto profile",
     "—",
     "1. Klik upload foto\n2. Pilih gambar\n3. Crop dengan react-cropper\n4. Save",
     "Foto ter-upload ke ImageKit, preview ter-update", "Medium"),

    # ===== H. ADMIN DASHBOARD =====
    ("ADM-001", "Admin", "UI/Functional", "Dashboard Home", "Lihat dashboard admin",
     "Login sebagai admin",
     "1. Buka /admin/dashboards",
     "Dashboard tampil: stats, recent activity, charts", "Medium"),
    ("ADM-002", "Admin", "UI/Functional", "User Management", "CRUD user (admin only)",
     "Login sebagai admin",
     "1. Buka /admin/dashboard/user\n2. Create/Edit/Delete user\n3. Assign role",
     "User ter-CRUD, role ter-assign", "High"),
    ("ADM-003", "Admin", "UI/Functional", "Campaign Management", "CRUD campaign donasi (admin+bismar)",
     "Login sebagai admin/bismar",
     "1. Buka /admin/dashboard/campaign\n2. CRUD campaign",
     "Campaign ter-CRUD", "High"),
    ("ADM-004", "Admin", "UI/Functional", "Campaign Categories", "Manage kategori campaign",
     "—",
     "1. Buka /admin/dashboard/category/campaign\n2. CRUD kategori",
     "Kategori ter-CRUD", "Medium"),
    ("ADM-005", "Admin", "UI/Functional", "Donation History", "Lihat history donasi per campaign",
     "—",
     "1. Buka /admin/dashboard/donation/[slug]",
     "List donasi per campaign tampil", "High"),
    ("ADM-006", "Admin", "UI/Functional", "Proposal Management", "CRUD proposal (admin only)",
     "—",
     "1. Buka /admin/dashboard/proposal\n2. CRUD proposal",
     "Proposal ter-CRUD", "Low"),
    ("ADM-007", "Admin", "Security", "Role Guard", "User tanpa role admin akses /admin/dashboard/user",
     "Login sebagai bismar",
     "1. Akses /admin/dashboard/user",
     "403 Forbidden atau redirect", "High"),
]

# ============================================================
# API TEST CASES
# ============================================================
# Format: (ID, Endpoint, Method, Tipe, Test Case, Pre-condition, Request, Expected, Priority)
API_TEST_CASES = [
    # ===== AUTH =====
    ("API-AUTH-001", "/auth/login", "POST", "Functional", "Login dengan kredensial valid",
     "User registered & verified",
     '{"email": "user@test.com", "password": "Valid123!"}',
     '200 OK, response: {token, user: {id, email, role}}', "High"),
    ("API-AUTH-002", "/auth/login", "POST", "Validation", "Login dengan email tidak valid",
     "—",
     '{"email": "bukan-email", "password": "abc"}',
     '400 Bad Request, errors.email', "Medium"),
    ("API-AUTH-003", "/auth/login", "POST", "Functional", "Login dengan kredensial salah",
     "—",
     '{"email": "user@test.com", "password": "wrong"}',
     '401 Unauthorized, message: "Invalid credentials"', "High"),
    ("API-AUTH-004", "/auth/register", "POST", "Functional", "Register user baru",
     "Email belum terdaftar",
     '{"name": "Test", "email": "new@test.com", "password": "Valid123!"}',
     '201 Created, user dibuat, email verifikasi dikirim', "High"),
    ("API-AUTH-005", "/auth/register", "POST", "Validation", "Register dengan email duplikat",
     "Email sudah ada",
     '{"email": "existing@test.com", ...}',
     '409 Conflict, message: "Email already registered"', "Medium"),
    ("API-AUTH-006", "/auth/logout", "POST", "Functional", "Logout",
     "User logged in",
     "(with valid token in header)",
     '200 OK, token invalidated server-side', "High"),
    ("API-AUTH-007", "/auth/me", "GET", "Functional", "Get current user",
     "User logged in",
     "(with valid token)",
     '200 OK, response: {id, name, email, role, avatar}', "High"),
    ("API-AUTH-008", "/auth/me", "GET", "Security", "Get current user tanpa token",
     "—",
     "(no token)",
     '401 Unauthorized', "High"),
    ("API-AUTH-009", "/auth/me", "GET", "Security", "Get current user dengan token expired",
     "—",
     "(expired token)",
     '401 Unauthorized, message: "Token expired"', "High"),
    ("API-AUTH-010", "/auth/forgot-password", "POST", "Functional", "Request reset password",
     "—",
     '{"email": "user@test.com"}',
     '200 OK, email reset terkirim', "Medium"),
    ("API-AUTH-011", "/auth/reset-password", "POST", "Functional", "Reset password dengan token",
     "Token dari email",
     '{"token": "...", "new_password": "New123!"}',
     '200 OK, password updated', "High"),
    ("API-AUTH-012", "/auth/verify-email", "POST", "Functional", "Verifikasi email",
     "—",
     '{"token": "verify_token_from_email"}',
     '200 OK, user ter-verifikasi', "High"),
    ("API-AUTH-013", "/auth/google", "POST", "Integration", "Google OAuth login",
     "—",
     '{"google_token": "..."}',
     '200 OK, user ter-login', "Medium"),
    ("API-AUTH-014", "/auth/refresh", "POST", "Functional", "Refresh JWT token",
     "—",
     '{"refresh_token": "..."}',
     '200 OK, new access_token', "High"),
    ("API-AUTH-015", "/auth/login", "POST", "Security", "Rate limiting (10x salah password)",
     "—",
     "10x POST dengan password salah dalam 5 menit",
     '429 Too Many Requests setelah 5x attempt', "Medium"),

    # ===== USER/PROFILE =====
    ("API-USER-001", "/user/profile", "GET", "Functional", "Get user profile",
     "User logged in",
     "(with valid token)",
     '200 OK, response: {id, name, email, role, avatar, ...}', "Medium"),
    ("API-USER-002", "/user/profile", "PUT", "Functional", "Update user profile",
     "—",
     '{"name": "Updated Name", "phone": "08123456789"}',
     '200 OK, profile updated', "Medium"),
    ("API-USER-003", "/user/avatar", "POST", "Integration", "Upload avatar ke ImageKit",
     "—",
     "multipart/form-data: file=@avatar.jpg",
     '200 OK, response: {avatar_url}', "Medium"),
    ("API-USER-004", "/user/profile", "GET", "Security", "Get profile user lain (bukan diri sendiri)",
     "—",
     "(with valid token, requesting other user)",
     '403 Forbidden', "Low"),

    # ===== BLOG =====
    ("API-BLOG-001", "/blog", "GET", "Functional", "List blog articles",
     "—",
     "?page=1&per_page=10",
     '200 OK, response: {data: [...], pagination}', "Medium"),
    ("API-BLOG-002", "/blog", "GET", "Functional", "List blog dengan search",
     "—",
     "?search=keyword",
     '200 OK, filtered results', "Medium"),
    ("API-BLOG-003", "/blog", "GET", "Functional", "List blog dengan category filter",
     "—",
     "?category=1",
     '200 OK, filtered by category', "Low"),
    ("API-BLOG-004", "/blog/:slug", "GET", "Functional", "Get single blog post",
     "—",
     "/blog/my-article",
     '200 OK, response: {id, title, content, cover_image, author, ...}', "High"),
    ("API-BLOG-005", "/blog/:slug", "GET", "Edge", "Get blog post yang tidak ada",
     "—",
     "/blog/not-exist",
     '404 Not Found', "Medium"),
    ("API-BLOG-006", "/blog", "POST", "Functional", "Create blog (admin/copywriter)",
     "Login as admin/copywriter",
     '{"title": "...", "content": "...", "category_id": 1}',
     '201 Created, blog created', "High"),
    ("API-BLOG-007", "/blog/:slug", "PUT", "Functional", "Update blog",
     "—",
     '{"title": "Updated", "content": "..."}',
     '200 OK, blog updated', "High"),
    ("API-BLOG-008", "/blog/:slug", "DELETE", "Functional", "Delete blog",
     "—",
     "(no body)",
     '200 OK, blog soft-deleted', "High"),
    ("API-BLOG-009", "/blog", "POST", "Security", "Create blog tanpa auth (role copywriter only)",
     "Login as bismar (no access)",
     '{"title": "...", "content": "..."}',
     '403 Forbidden', "High"),
    ("API-BLOG-010", "/blog/categories", "GET", "Functional", "List blog categories",
     "—",
     "(no params)",
     '200 OK, response: {data: [{id, name}, ...]}', "Low"),
    ("API-BLOG-011", "/blog/upload-image", "POST", "Integration", "Upload image ke ImageKit",
     "—",
     "multipart: file=@image.jpg",
     '200 OK, response: {image_url}', "High"),

    # ===== CAMPAIGN / DONATION =====
    ("API-CAMP-001", "/campaign", "GET", "Functional", "List campaigns",
     "—",
     "?page=1&per_page=10",
     '200 OK, list campaigns with target/progress', "High"),
    ("API-CAMP-002", "/campaign/:slug", "GET", "Functional", "Get single campaign",
     "—",
     "/campaign/bantu-anak",
     '200 OK, detail campaign', "High"),
    ("API-CAMP-003", "/campaign", "POST", "Functional", "Create campaign (admin/bismar)",
     "—",
     '{"title": "...", "target": 10000000, "deadline": "2026-12-31"}',
     '201 Created', "High"),
    ("API-CAMP-004", "/campaign/:id", "PUT", "Functional", "Update campaign",
     "—",
     '{"target": 15000000}',
     '200 OK, campaign updated', "High"),
    ("API-CAMP-005", "/campaign/:id", "DELETE", "Functional", "Delete campaign",
     "—",
     "(no body)",
     '200 OK, campaign deleted', "High"),
    ("API-CAMP-006", "/donation", "POST", "Critical", "Create donation (start transaction)",
     "User logged in",
     '{"campaign_id": 1, "amount": 100000, "donor_name": "Anon"}',
     '201 Created, response: {donation_id, snap_token}', "High"),
    ("API-CAMP-007", "/donation/:id", "GET", "Functional", "Get donation detail",
     "—",
     "(with valid id)",
     '200 OK, donation detail', "Medium"),
    ("API-CAMP-008", "/midtrans/callback", "POST", "Critical", "Midtrans payment callback (success)",
     "Donation created, payment success",
     '{"transaction_id": "...", "status": "settlement"}',
     '200 OK, donation updated to "paid"', "High"),
    ("API-CAMP-009", "/midtrans/callback", "POST", "Critical", "Midtrans payment callback (failure)",
     "Donation created, payment failed",
     '{"transaction_id": "...", "status": "deny"}',
     '200 OK, donation updated to "failed"', "High"),
    ("API-CAMP-010", "/midtrans/callback", "POST", "Security", "Midtrans callback dengan signature invalid",
     "—",
     '{"transaction_id": "...", "signature_key": "fake"}',
     '401 Unauthorized', "High"),
    ("API-CAMP-011", "/donation/history", "GET", "Functional", "Get user donation history",
     "User logged in",
     "(with valid token)",
     '200 OK, list donations by current user', "Medium"),
    ("API-CAMP-012", "/admin/donations", "GET", "Functional", "Admin get all donations",
     "Login as admin",
     "?campaign_id=1",
     '200 OK, all donations', "High"),

    # ===== EVENT =====
    ("API-EVT-001", "/event", "GET", "Functional", "List public events",
     "—",
     "?page=1",
     '200 OK, list events', "Medium"),
    ("API-EVT-002", "/event/:slug", "GET", "Functional", "Get event detail",
     "—",
     "/event/webinar-xyz",
     '200 OK, event detail', "Medium"),
    ("API-EVT-003", "/event/:id/register", "POST", "Functional", "Register for event",
     "User logged in",
     '{"event_id": 1}',
     '201 Created, registration success', "High"),
    ("API-EVT-004", "/event/:id/register", "POST", "Edge", "Register event yang sudah lewat",
     "—",
     '{"event_id": 999}',  # past event
     '400 Bad Request, "Event sudah selesai"', "Medium"),
    ("API-EVT-005", "/admin/event", "POST", "Functional", "Create event (admin only)",
     "Login as admin",
     '{"title": "...", "date": "2026-12-01", "location": "..."}',
     '201 Created', "High"),
    ("API-EVT-006", "/admin/event/:id", "PUT", "Functional", "Update event",
     "—",
     '{"title": "Updated", "date": "..."}',
     '200 OK', "Medium"),
    ("API-EVT-007", "/admin/event/:id", "DELETE", "Functional", "Delete event",
     "—",
     "(no body)",
     '200 OK', "Medium"),

    # ===== VOLUNTEER =====
    ("API-VOL-001", "/volunteer/register", "POST", "Functional", "Daftar volunteer",
     "—",
     '{"name": "...", "email": "...", "phone": "...", "skills": [...]}',
     '201 Created, registration success', "High"),
    ("API-VOL-002", "/admin/volunteers", "GET", "Functional", "List volunteers (admin)",
     "Login as admin",
     "?status=pending",
     '200 OK, list volunteers', "High"),
    ("API-VOL-003", "/admin/volunteers/:id/approve", "POST", "Functional", "Approve volunteer",
     "—",
     "(no body)",
     '200 OK, status updated to "approved"', "High"),
    ("API-VOL-004", "/admin/volunteers/:id/reject", "POST", "Functional", "Reject volunteer",
     "—",
     '{"reason": "..."}',
     '200 OK, status updated to "rejected"', "High"),

    # ===== USER MANAGEMENT (admin) =====
    ("API-ADM-001", "/admin/users", "GET", "Functional", "List users (admin)",
     "Login as admin",
     "?page=1",
     '200 OK, list users with role', "High"),
    ("API-ADM-002", "/admin/users", "POST", "Functional", "Create user (admin)",
     "—",
     '{"name": "...", "email": "...", "role": "bismar", "password": "..."}',
     '201 Created', "High"),
    ("API-ADM-003", "/admin/users/:id", "PUT", "Functional", "Update user role",
     "—",
     '{"role": "copywriter"}',
     '200 OK', "High"),
    ("API-ADM-004", "/admin/users/:id", "DELETE", "Functional", "Delete user",
     "—",
     "(no body)",
     '200 OK', "High"),
    ("API-ADM-005", "/admin/users", "GET", "Security", "List users tanpa role admin",
     "Login as bismar",
     "(with bismar token)",
     '403 Forbidden', "High"),
]

# ============================================================
# GENERATE XLSX
# ============================================================
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# === SHEET 1: Test Cases (UI/Functional) ===
ws1 = wb.create_sheet("Test Cases")
headers1 = ["ID", "Modul", "Tipe", "Item", "Test Case", "Pre-condition", "Steps", "Expected Result", "Priority", "Status", "Actual Result", "Catatan"]
ws1.append(headers1)

for tc in TEST_CASES:
    # Status default "Belum Dieksekusi", Actual & Catatan kosong
    row = list(tc) + ["Belum Dieksekusi", "", ""]
    ws1.append(row)

# Style header
for col in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

# Style data rows
for row_idx in range(2, len(TEST_CASES) + 2):
    for col_idx in range(1, len(headers1) + 1):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER

    # Color priority column (col 9)
    priority = ws1.cell(row=row_idx, column=9).value
    if priority == "High":
        ws1.cell(row=row_idx, column=9).fill = PRIORITY_HIGH
    elif priority == "Medium":
        ws1.cell(row=row_idx, column=9).fill = PRIORITY_MED
    elif priority == "Low":
        ws1.cell(row=row_idx, column=9).fill = PRIORITY_LOW

# Column widths
col_widths1 = [12, 12, 18, 22, 40, 30, 50, 50, 10, 16, 30, 25]
for i, w in enumerate(col_widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Row heights
ws1.row_dimensions[1].height = 30

# Freeze header
ws1.freeze_panes = "A2"

# Auto filter
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{len(TEST_CASES) + 1}"

# === SHEET 2: API Test Cases ===
ws2 = wb.create_sheet("API Test Cases")
headers2 = ["ID", "Endpoint", "Method", "Tipe", "Test Case", "Pre-condition", "Request Body / Params", "Expected Response", "Priority", "Status", "Actual Result", "Catatan"]
ws2.append(headers2)

for tc in API_TEST_CASES:
    row = list(tc) + ["Belum Dieksekusi", "", ""]
    ws2.append(row)

# Style
for col in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=col)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

for row_idx in range(2, len(API_TEST_CASES) + 2):
    for col_idx in range(1, len(headers2) + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER

    priority = ws2.cell(row=row_idx, column=9).value
    if priority == "High":
        ws2.cell(row=row_idx, column=9).fill = PRIORITY_HIGH
    elif priority == "Medium":
        ws2.cell(row=row_idx, column=9).fill = PRIORITY_MED
    elif priority == "Low":
        ws2.cell(row=row_idx, column=9).fill = PRIORITY_LOW

col_widths2 = [14, 28, 8, 18, 40, 30, 40, 50, 10, 16, 30, 25]
for i, w in enumerate(col_widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(API_TEST_CASES) + 1}"

# === SHEET 3: Summary ===
ws3 = wb.create_sheet("Summary", 0)  # insert as first sheet
ws3.append(["IFL Chapter Malang — Test Case Summary"])
ws3.append([])
ws3.append(["Document Info", ""])
ws3.append(["Project", "IFL Chapter Malang (website-ifl)"])
ws3.append(["Source Repo", "github.com/irzadzulhika29/fe_iflchaptermalang"])
ws3.append(["Test Plan", "test-plan.md (BINUS framework, 19 sections)"])
ws3.append(["Generated", "2026-06-22"])
ws3.append([])
ws3.append(["Test Inventory", "Count"])
ws3.append(["UI Test Cases", len(TEST_CASES)])
ws3.append(["API Test Cases", len(API_TEST_CASES)])
ws3.append(["Total", len(TEST_CASES) + len(API_TEST_CASES)])
ws3.append([])

# Per-modul count
ws3.append(["By Modul (UI)", "Count"])
modul_count = {}
for tc in TEST_CASES:
    modul = tc[1]
    modul_count[modul] = modul_count.get(modul, 0) + 1
for modul, count in sorted(modul_count.items()):
    ws3.append([f"  {modul}", count])

ws3.append([])

# Per-priority
ws3.append(["By Priority (UI)", "Count"])
prio_count = {"High": 0, "Medium": 0, "Low": 0}
for tc in TEST_CASES:
    prio_count[tc[8]] = prio_count.get(tc[8], 0) + 1
for prio, count in prio_count.items():
    ws3.append([f"  {prio}", count])

ws3.append([])

ws3.append(["By Endpoint (API)", "Count"])
ep_count = {}
for tc in API_TEST_CASES:
    # Group by base endpoint
    endpoint = tc[1].split("/")[1] if "/" in tc[1] else tc[1]
    ep_count[endpoint] = ep_count.get(endpoint, 0) + 1
for ep, count in sorted(ep_count.items()):
    ws3.append([f"  /{ep}", count])

# Style summary
ws3.cell(row=1, column=1).font = Font(size=14, bold=True, color="1E40AF")
ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor="DBEAFE")

for r in [3, 10, 17, 24]:
    cell = ws3.cell(row=r, column=1)
    cell.font = Font(bold=True, color="1E40AF")

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 15

# === SAVE ===
wb.save(OUTPUT)
size = os.path.getsize(OUTPUT)
print(f"✓ XLSX saved: {OUTPUT}")
print(f"  Size: {size:,} bytes ({size/1024:.1f} KB)")
print(f"  Sheets: 3 (Summary, Test Cases, API Test Cases)")
print(f"  Test Cases (UI): {len(TEST_CASES)}")
print(f"  Test Cases (API): {len(API_TEST_CASES)}")
print(f"  Total: {len(TEST_CASES) + len(API_TEST_CASES)}")
