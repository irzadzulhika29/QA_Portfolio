"""Tambah negative + edge test cases ke IFL XLSX."""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = "/home/ubuntu/testing-hub/projects/ifl-chapter-malang/test-cases.xlsx"

# Styling
DATA_FONT = Font(name="Arial", size=10)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
PRIORITY_HIGH = PatternFill("solid", fgColor="FECACA")
PRIORITY_MED = PatternFill("solid", fgColor="FED7AA")
PRIORITY_LOW = PatternFill("solid", fgColor="DCFCE7")
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 25 Negative + 5 Edge UI test cases
NEW_UI = [
    # Auth negative
    ("AUTH-018", "Auth", "Negative", "Login",
     "Login dengan email mengandung SQL injection",
     "—",
     "1. Buka /masuk\n2. Isi email: ' OR 1=1--\n3. Password apa saja\n4. Klik Masuk",
     "Login ditolak, error 'Format email tidak valid' (input disanitized)", "High"),
    ("AUTH-019", "Auth", "Negative", "Login",
     "Login dengan password mengandung XSS",
     "—",
     "1. Isi password: <script>alert(1)</script>\n2. Submit",
     "Input di-escape, no script execution, login tetap proper", "High"),
    ("AUTH-020", "Auth", "Negative", "Register",
     "Register dengan password terlalu pendek",
     "—",
     "1. Buka /daftar\n2. Password: 'abc'\n3. Submit",
     "Validasi form gagal, error 'Password minimal 8 karakter'", "Medium"),
    ("AUTH-021", "Auth", "Negative", "Google OAuth",
     "Google OAuth callback dari provider tidak dikenal",
     "—",
     "1. Mock callback dari domain bukan accounts.google.com",
     "401 Unauthorized, token google invalid", "High"),
    ("AUTH-022", "Auth", "Negative", "ProtectedRoute",
     "User tanpa role akses dashboard (cookie tampering)",
     "—",
     "1. Modify cookie role jadi 'admin'\n2. Buka /admin/dashboards",
     "Backend verify role via JWT, 403 Forbidden", "High"),
    ("AUTH-023", "Auth", "Negative", "Login",
     "Brute force: 10x login gagal bergantian",
     "—",
     "1. Submit password salah 10x dalam 5 menit",
     "Setelah 5x: 429 Too Many Requests", "High"),

    # Donasi negative
    ("DON-017", "Donasi", "Negative", "Donasi",
     "Donasi dengan nominal di bawah minimum",
     "Chatbot terbuka",
     "1. Input '100' (< min 10000)",
     "Validation feedback: 'Nominal minimum Rp 10.000'", "Medium"),
    ("DON-018", "Donasi", "Negative", "Donasi",
     "Donasi dengan nominal di atas maksimum (1 M)",
     "—",
     "1. Input '10000000000' (10M, over max)",
     "Validation: 'Nominal maksimal Rp 10.000.000'", "Medium"),
    ("DON-019", "Donasi", "Negative", "Midtrans",
     "Midtrans Snap gagal di-load (adblock)",
     "—",
     "1. Blokir midtrans.com di adblocker\n2. Klik Bayar",
     "Loading state tampil, timeout setelah 30 detik, error message, bisa coba lagi", "High"),
    ("DON-020", "Donasi", "Negative", "Midtrans",
     "Midtrans callback signature invalid (tampered)",
     "—",
     "1. Manipulasi Midtrans callback signature_key",
     "401 Unauthorized, callback rejected", "High"),
    ("DON-021", "Donasi", "Negative", "QRIS",
     "QRIS payment expired sebelum konfirmasi",
     "—",
     "1. Minta QRIS\n2. Tunggu 15 menit tidak bayar",
     "Payment expired, kembali ke summary, bisa coba metode lain", "Medium"),
    ("DON-022", "Donasi", "Negative", "Double Submit",
     "Double click bayar (race condition)",
     "Di summary",
     "1. Klik Bayar 2x cepat",
     "Hanya 1 Snap yang terbuat, idempotent (transaction_id tidak duplikat)", "High"),
    ("DON-023", "Donasi", "Negative", "Campaign",
     "Donasi ke campaign yang sudah closed (target tercapai)",
     "—",
     "1. Buka campaign yang sudah 100%\n2. Klik Donasi",
     "Campaign marked 'Terkumpul', tombol donasi disabled atau ada warning", "High"),

    # Blog negative
    ("BLG-012", "Blog", "Negative", "CKEditor",
     "CKEditor XSS injection di konten blog",
     "Admin login",
     "1. Add blog\n2. Konten: <script>alert('xss')</script>\n3. Publish",
     "Script termakanerakan (di-sanitize) di frontend, tidak execute di browser", "High"),
    ("BLG-013", "Blog", "Negative", "ImageKit",
     "Upload file non-image (exe/pdf) sebagai cover",
     "—",
     "1. Klik upload cover\n2. Pilih file .exe atau .pdf",
     "Filter tipe file, error 'Hanya gambar yang diizinkan'", "Medium"),
    ("BLG-014", "Blog", "Negative", "ImageKit",
     "Upload gambar > 5MB",
     "—",
     "1. Upload gambar 10MB",
     "Error: 'Ukuran file maksimal 5MB'", "Medium"),
    ("BLG-015", "Blog", "Negative", "Blog",
     "Hapus blog yang sudah di-publish (tanpa draft dulu)",
     "—",
     "1. Klik Delete di blog yang sudah live",
     "Konfirmasi: 'Apakah Anda yakin ingin menghapus blog ini?' + warning", "Medium"),

    # Event negative
    ("EVT-006", "Event", "Negative", "Event",
     "Daftar event yang sudah lewat",
     "—",
     "1. Buka event past deadline\n2. Klik Daftar",
     "Error: 'Pendaftaran sudah ditutup'", "Medium"),
    ("EVT-007", "Event", "Negative", "Event",
     "Daftar event 2x (duplicate registration)",
     "—",
     "1. Daftar event\n2. Daftar lagi di event sama",
     "Error: 'Anda sudah terdaftar di event ini'", "Medium"),

    # Volunteer negative
    ("VOL-006", "Volunteer", "Negative", "Volunteer",
     "Register volunteer tanpa nama (empty)",
     "—",
     "1. Isi form volunteer, kosongkan nama\n2. Submit",
     "Validasi: 'Nama wajib diisi'", "Medium"),
    ("VOL-007", "Volunteer", "Negative", "Volunteer",
     "Register volunteer dengan email tidak valid",
     "—",
     "1. Isi email: 'bukan-email'\n2. Submit",
     "Validasi: 'Format email tidak valid'", "Medium"),

    # Profile negative
    ("PROF-005", "Profile", "Negative", "Profile",
     "Upload profile pic file berbahaya (.exe)",
     "—",
     "1. Klik upload foto\n2. Pilih file .exe",
     "Filter: 'Hanya file gambar yang diizinkan'", "Medium"),
    ("PROF-006", "Profile", "Negative", "Profile",
     "Edit profile dengan bio > 500 karakter",
     "—",
     "1. Edit bio 1000 karakter\n2. Save",
     "Validasi: 'Bio maksimal 500 karakter'", "Low"),

    # Admin dashboard negative
    ("ADM-008", "Admin", "Negative", "Dashboard",
     "Akses admin endpoint langsung tanpa login",
     "—",
     "1. Buka /admin/dashboards langsung (tanpa session)",
     "Redirect ke /masuk, query param ?redirect=/admin/dashboards", "High"),
    ("ADM-009", "Admin", "Security", "Role Guard",
     "User role 'bismar' akses admin-only endpoint",
     "Login sebagai bismar",
     "1. Buka /admin/dashboard/user (admin only)",
     "403 Forbidden atau redirect", "High"),
    ("ADM-010", "Admin", "Security", "Role Guard",
     "User role 'copywriter' akses event management",
     "Login sebagai copywriter",
     "1. Buka /admin/dashboard/event (admin+bismar only)",
     "403 Forbidden atau redirect", "High"),

    # Public edge
    ("PUB-013", "Public", "Edge", "404",
     "Akses URL dengan karakter spesial (unicode, ~, $)",
     "—",
     "1. Buka /test/~$%25/xyz",
     "404 halaman tampil tanpa error (tidak crash)", "Low"),
    ("PUB-014", "Public", "Edge", "Responsive",
     "Responsive di iPad landscape (1024x768)",
     "—",
     "1. Buka / di viewport 1024x768",
     "Layout responsive, tidak ada horizontal scroll", "Medium"),
]

# Load & append
wb = load_workbook(XLSX)
ws = wb["Test Cases"]
next_row = ws.max_row + 1

for tc in NEW_UI:
    row_data = list(tc) + ["Belum Dieksekusi", "", ""]
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=next_row, column=col_idx, value=val)
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=next_row, column=col_idx)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGN
        cell.border = BORDER
    if tc[8] == "High":
        ws.cell(row=next_row, column=9).fill = PRIORITY_HIGH
    elif tc[8] == "Medium":
        ws.cell(row=next_row, column=9).fill = PRIORITY_MED
    elif tc[8] == "Low":
        ws.cell(row=next_row, column=9).fill = PRIORITY_LOW
    next_row += 1

# Update Summary
ws_sum = wb["Summary"]
ui_total = ws.max_row - 1
# Find "UI Test Cases" row
for r in range(10, 20):
    v = ws_sum.cell(row=r, column=1).value
    if v and v.strip() == "UI Test Cases":
        ws_sum.cell(row=r, column=2, value=ui_total)
    elif v and v.strip() == "Total":
        # Find API count
        for r2 in range(10, 20):
            v2 = ws_sum.cell(row=r2, column=1).value
            if v2 and v2.strip() == "API Test Cases":
                api_total = ws_sum.cell(row=r2, column=2).value or 0
                ws_sum.cell(row=r, column=2, value=ui_total + api_total)
                break

# Update by-modul
modul_added = {}
for tc in NEW_UI:
    m = tc[1]
    modul_added[m] = modul_added.get(m, 0) + 1

for r in range(15, 35):
    ca = ws_sum.cell(row=r, column=1).value
    if ca is None or not isinstance(ca, str): continue
    ca_stripped = ca.strip()
    for m in modul_added:
        if ca_stripped == m:
            old = ws_sum.cell(row=r, column=2).value
            try:
                old_int = int(old) if old is not None else 0
            except (TypeError, ValueError):
                old_int = 0
            ws_sum.cell(row=r, column=2, value=old_int + modul_added[m])
            break

# Update by-priority
prio_added = {"High": 0, "Medium": 0, "Low": 0}
for tc in NEW_UI:
    p = tc[8]
    if p in prio_added:
        prio_added[p] += 1

for r in range(20, 35):
    ca = ws_sum.cell(row=r, column=1).value
    if ca is None or not isinstance(ca, str): continue
    ca_stripped = ca.strip()
    for p in prio_added:
        if ca_stripped == p:
            old = ws_sum.cell(row=r, column=2).value
            try:
                old_int = int(old) if old is not None else 0
            except:
                old_int = 0
            ws_sum.cell(row=r, column=2, value=old_int + prio_added[p])
            break

wb.save(XLSX)
size = os.path.getsize(XLSX)
print(f"OK IFL XLSX updated: {XLSX}")
print(f"  Size: {size:,} bytes")
print(f"  UI Test Cases: {ui_total} (added {len(NEW_UI)})")
print(f"\nAdded: {len(NEW_UI)} test cases")
print(f"  Auth: +6, Donasi: +7, Blog: +4, Event: +2, Volunteer: +2, Profile: +2, Admin: +3, Public: +2")
